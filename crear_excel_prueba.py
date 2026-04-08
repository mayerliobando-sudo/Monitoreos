"""
Script para crear un archivo Excel de prueba con la misma estructura
que lee el generador_dashboard.py (monitoreos 2025.xlsx).

Estructura por hoja (una hoja por mes):
- Columnas: Monitoreo Fecha | Horario Control | Aplicativo / Sistema | Inconvenientes | Comentario Admin
- Los meses van de Mayo a Diciembre 2025.
- Se generan datos realistas de alarmas para sistemas de la PGN.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

# ============================================================
# CONFIGURACION
# ============================================================
ARCHIVO_SALIDA = 'monitoreos_prueba_2025.xlsx'

SISTEMAS = [
    'ALFA', 'APP MOVIL', 'APPS EXTERNAS', 'APPS INTRANET', 'APPS PORTAL',
    'GESTOR DOKUS', 'HOMINIS', 'IGA', 'INSAP', 'ITA',
    'NUEVA SEDE ELECTRONICA', 'PORTAL WEB E INTRANET', 'SIAF',
    'SIGDEA PORTAL EMPLEADO', 'SIGDEA SEDE ELECTRONICA', 'SIM',
    'SIRI', 'STRATEGOS', 'X-ROAD'
]

MESES = {
    'Mayo':       (datetime(2025, 5, 1),  datetime(2025, 5, 31)),
    'Junio':      (datetime(2025, 6, 1),  datetime(2025, 6, 30)),
    'Julio':      (datetime(2025, 7, 1),  datetime(2025, 7, 31)),
    'Agosto':     (datetime(2025, 8, 1),  datetime(2025, 8, 31)),
    'Septiembre': (datetime(2025, 9, 1),  datetime(2025, 9, 30)),
    'Octubre':    (datetime(2025, 10, 1), datetime(2025, 10, 31)),
    'Noviembre':  (datetime(2025, 11, 1), datetime(2025, 11, 30)),
    'Diciembre':  (datetime(2025, 12, 1), datetime(2025, 12, 31)),
}

HORARIOS = ['Mañana', 'Medio dia', 'Tarde']

INCONVENIENTES = [
    'El sistema presenta lentitud extrema al cargar modulos principales',
    'Error de conexion a la base de datos, servicio caido',
    'Pantalla en rojo al intentar acceder al login',
    'Fallo intermitente en el modulo de consultas',
    'Timeout en la conexion al servidor de aplicaciones',
    'El sistema no responde a peticiones de usuarios',
    'Error 503 - Servicio no disponible temporalmente',
    'Caida total del servicio por mantenimiento no programado',
    'Error de autenticacion masiva en el directorio activo',
    'Problemas de rendimiento por alta demanda de usuarios',
    'El modulo de reportes genera error al exportar datos',
    'Fallo en la sincronizacion con el sistema central',
    'Interrupcion del servicio por fallo en el servidor principal',
    'Error critico en la base de datos del aplicativo',
    'Demora excesiva en el procesamiento de solicitudes',
]

ESTADOS_OK = ['OK', 'O.K.', 'Sin novedad', 'Ninguno', 'NA']

COMENTARIOS_ADMIN = [
    'Se escalo al proveedor. Ticket #2025-{ticket}',
    'Se reinicio el servicio y quedo operativo',
    'En espera de respuesta del equipo de infraestructura',
    'Se aplico parche temporal, pendiente solucion definitiva',
    'Se notifico al area de TI para revision urgente',
    'Se realizo reinicio del servidor, servicio restaurado',
    'Problema recurrente, se programo mantenimiento preventivo',
    'Se contacto al proveedor del hosting externo',
    'Se activo el plan de contingencia del servicio',
    'Pendiente revision por parte del administrador de BD',
]

# ============================================================
# GENERAR DATOS
# ============================================================
random.seed(42)  # Para reproducibilidad

wb = openpyxl.Workbook()
# Eliminar la hoja por defecto
wb.remove(wb.active)

# Estilos
azul_header = PatternFill(start_color='003876', end_color='003876', fill_type='solid')
amarillo_fila = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
blanco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
font_normal = Font(name='Segoe UI', size=10, color='1a1a1a')
borde = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

COLUMNAS = [
    'Monitoreo Fecha',
    'Horario Control',
    'Aplicativo / Sistema',
    'Inconvenientes',
    'Comentario Admin'
]
ANCHOS = [18, 18, 28, 45, 45]

for mes_nombre, (fecha_inicio, fecha_fin) in MESES.items():
    ws = wb.create_sheet(title=mes_nombre)

    # Encabezados
    for col_idx, nombre in enumerate(COLUMNAS, 1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font = font_header
        celda.fill = azul_header
        celda.alignment = center
        celda.border = borde

    # Generar entre 15 y 40 registros por mes
    # Mezcla de registros OK y alarmas
    num_registros = random.randint(20, 45)
    fila_actual = 2

    # Generar fechas aleatorias dentro del mes
    dias_en_mes = (fecha_fin - fecha_inicio).days + 1
    fechas_monitoreo = sorted(random.choices(
        [fecha_inicio + timedelta(days=d) for d in range(dias_en_mes)],
        k=num_registros
    ))

    fecha_anterior = None
    for i in range(num_registros):
        fecha = fechas_monitoreo[i]
        horario = random.choice(HORARIOS)

        # 60% de probabilidad de ser una alarma, 40% OK
        es_alarma = random.random() < 0.60

        if es_alarma:
            sistema = random.choice(SISTEMAS)
            inconveniente = random.choice(INCONVENIENTES)
            ticket = random.randint(1000, 9999)
            comentario = random.choice(COMENTARIOS_ADMIN).format(ticket=ticket)
        else:
            sistema = random.choice(SISTEMAS)
            inconveniente = random.choice(ESTADOS_OK)
            comentario = ''

        # Escribir la fecha como texto en formato "dd de mes de aaaa"
        meses_texto = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
            5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
            9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }
        fecha_texto = f"{fecha.day} de {meses_texto[fecha.month]} de {fecha.year}"

        # Solo escribir la fecha si cambio (forward fill logic)
        if fecha != fecha_anterior:
            ws.cell(row=fila_actual, column=1, value=fecha_texto).font = font_normal
        fecha_anterior = fecha

        ws.cell(row=fila_actual, column=2, value=horario).font = font_normal
        ws.cell(row=fila_actual, column=3, value=sistema).font = font_normal
        ws.cell(row=fila_actual, column=4, value=inconveniente).font = font_normal
        ws.cell(row=fila_actual, column=5, value=comentario).font = font_normal

        # Aplicar estilos
        fill = amarillo_fila if (fila_actual % 2 == 0) else blanco
        for col_idx in range(1, 6):
            ws.cell(row=fila_actual, column=col_idx).fill = fill
            ws.cell(row=fila_actual, column=col_idx).border = borde
            if col_idx <= 3:
                ws.cell(row=fila_actual, column=col_idx).alignment = center
            else:
                ws.cell(row=fila_actual, column=col_idx).alignment = left_align

        fila_actual += 1

    # Ajustar anchos de columna
    for col_idx, ancho in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.sheet_properties.tabColor = '003876'

wb.save(ARCHIVO_SALIDA)
print(f"\n{'='*60}")
print(f"  ARCHIVO EXCEL DE PRUEBA CREADO EXITOSAMENTE")
print(f"{'='*60}")
print(f"  Archivo: {ARCHIVO_SALIDA}")
print(f"  Hojas:   {', '.join(MESES.keys())}")
print(f"  Estructura por hoja:")
print(f"    - Monitoreo Fecha (formato: 'dd de mes de aaaa')")
print(f"    - Horario Control (Mañana/Medio dia/Tarde)")
print(f"    - Aplicativo / Sistema ({len(SISTEMAS)} sistemas)")
print(f"    - Inconvenientes (alarmas o estados OK)")
print(f"    - Comentario Admin")
print(f"{'='*60}")
print(f"\nPara generar el dashboard con este archivo ejecuta:")
print(f"  python generador_dashboard.py {ARCHIVO_SALIDA}")
